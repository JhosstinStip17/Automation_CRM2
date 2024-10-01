"""Automation_CRM2"""

import os
import time
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException


class CRM2Automation:
    """Clase Automatización"""

    def __init__(self):
        """Metodo contructor de la clase"""

        # Se crea las opciones para abrir una instancia de google existente
        self.chrome_option = Options()
        # Se agrega la opcion para que use una instanacia de google exitente
        self.chrome_option.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
        # Se crea el driver
        self.driver = webdriver.Chrome(options=self.chrome_option)
        # Se define el tiempo de espera para el driver
        self.wait = WebDriverWait(self.driver, 10)
        # Variable para el manejo de nuevas sección
        self.current_section = 1
        # Variable para contar la cantidad de campos
        self.section_index = 0

    # Metodo para colocar las opciones en los campos que lo requieran
    def add_options(self, list_options):
        """Metodo para agregar las opciones a los deplegable"""

        # Bucle para colocar las opciones
        for i, option in enumerate(list_options):
            xpath_option = f"/html/body/app-root/app-mios/app-side-bar/div/mat-sidenav-container/mat-sidenav-content/div/app-admin-forms/div/div[2]/mat-tab-group/div/mat-tab-body[2]/div/div/div/div[1]/div[2]/div[{i+1 if 1 > 0 else ''}]/div[1]/mat-form-field[1]/div/div[1]/div/input"
            try:
                option_input = self.wait.until(
                    EC.presence_of_element_located((By.XPATH, xpath_option))
                )

                option_input.send_keys(option)

                # condicional para agregar un elemento opcion hasta que sea el ultimo
                if i < len(list_options) - 1:
                    button_add = self.driver.find_element(
                        By.XPATH,
                        "/html/body/app-root/app-mios/app-side-bar/div/mat-sidenav-container/mat-sidenav-content/div/app-admin-forms/div/div[2]/mat-tab-group/div/mat-tab-body[2]/div/div/div/div[1]/div[2]/div/div[2]/button",
                    )
                    button_add.click()
            except (TimeoutException, NoSuchElementException) as e:
                print(f"NO SE PUDO AGREGAR LA OPCION {option}. ERROR: {str(e)}")

    def read_user_from_excel(self, file_path, sheet_name, star_cell, end_cell):
        """Metodo para leer los usuarios de un archivo Excel"""

        # Variable para alamcenar el documento excel
        excel = load_workbook(filename=file_path)
        # Entrar en al hoja del excel
        sheet = excel[sheet_name]
        # Lista para alamcenar los usuarios
        users = []
        # Bucle para buscar a los usuarios y almacenarlos en la lista
        for row in sheet[star_cell:end_cell]:
            for cell in row:
                if cell.value is not None:
                    # Se coloca str para converti el valor en un string
                    users.append(str(cell.value))
        return users

    def create_group(self, campaign_name, group_name, group_descrip, users_to_add):
        """Metodo para crear el grupo"""

        # Varible modulo grupo
        group_button = self.driver.find_element(
            By.XPATH,
            "/html/body/app-root/app-mios/app-side-bar/div/mat-sidenav-container/mat-sidenav/div/app-left-nav/div/div/div/mat-nav-list/div/a[3]",
        )
        group_button.click()

        time.sleep(1)

        # Variable al boton crear grupo
        create_group = self.wait.until(
            EC.visibility_of_element_located(
                (
                    By.XPATH,
                    "/html/body/app-root/app-mios/app-side-bar/div/mat-sidenav-container/mat-sidenav-content/div/app-groups-list/div/div/div[1]/div[2]/div/button",
                )
            )
        )
        create_group.click()

        time.sleep(2)

        for a in range(3):
            campaigns = None
            try:
                # Variable a select de campañas dentro de las caracteristicas de crear grupo
                campaigns = self.driver.find_element(
                    By.XPATH,
                    f"/html/body/div[{a+2}]/div[2]/div/mat-dialog-container/app-admin-groups/form/mat-dialog-content/mat-form-field[1]/div/div[1]/div/mat-select",
                )
                break
            except (TimeoutException, NoSuchElementException):
                continue
        if campaigns:
            try:
                campaigns.click()
            except ImportError as e:
                print(f"NO SE PUDO DAR CLICK AL ELEMENTO 'CAMPAÑAS' ERROR : {e}")
        else:
            print("NO SE PUDO ENCONTRAR EL ELEMENTO 'CAMPAÑAS'")

        time.sleep(1.5)

        # Bucle para buscar la campaña por nombre
        for attempt in range(5):
            select_campaign = None
            try:
                select_campaign = self.driver.find_element(
                    By.XPATH, f"//mat-option[contains(., '{campaign_name}')]"
                )

                break
            except (TimeoutException, NoSuchElementException):
                continue
        if select_campaign:
            try:
                select_campaign.click()
                select_campaign.send_keys(Keys.TAB)
            except ImportError as e:
                print(f"NO SE PUDO ELEGIR LA CAMPAÑA: {campaign_name} ERRO :{e}")
        else:
            print(f"NO SE ENCONTRO LA CAMPAÑA: {campaign_name}")

        time.sleep(0.5)

        for a in range(4):
            try:
                # Coloca el nombre del grupo
                name_group = self.driver.find_element(
                    By.XPATH,
                    f"/html/body/div[{a+2}]/div[2]/div/mat-dialog-container/app-admin-groups/form/mat-dialog-content/mat-form-field[2]/div/div[1]/div/input",
                )

                name_group.send_keys(group_name)
                break
            except (TimeoutException, NoSuchElementException):
                continue

        for a in range(3):
            try:
                # Coloca la descripcion del grupo
                descrip_group = self.driver.find_element(
                    By.XPATH,
                    f"/html/body/div[{a+2}]/div[2]/div/mat-dialog-container/app-admin-groups/form/mat-dialog-content/mat-form-field[3]/div/div[1]/div/input",
                )

                descrip_group.send_keys(group_descrip)
                break
            except (TimeoutException, NoSuchElementException):
                continue

        # Bucle para añadir a los usuarios
        for user_to_add in users_to_add:
            button_add = None
            for a in range(3):
                try:
                    add_user = self.driver.find_element(
                        By.XPATH,
                        f"/html/body/div[{a+2}]/div[2]/div/mat-dialog-container/app-admin-groups/form/mat-dialog-content/div[4]/div[1]/mat-form-field/div/div[1]/div[1]/input",
                    )
                    add_user.clear()
                    add_user.send_keys(user_to_add)
                    add_user.send_keys(Keys.ENTER)

                    button_add = self.wait.until(
                        EC.element_to_be_clickable(
                            (
                                By.XPATH,
                                f"/html/body/div[{a+2}]/div[2]/div/mat-dialog-container/app-admin-groups/form/mat-dialog-content/div[4]/div[2]/button",
                            )
                        )
                    )
                    break
                except (TimeoutException, NoSuchElementException):
                    continue

            # Verificar si se ha encontrado el botón y se ha asignado
            if button_add:
                try:
                    # Intentar hacer clic en el botón para añadir el usuario
                    button_add.click()
                except Exception as e:
                    # Manejar cualquier error durante el clic y continuar con el siguiente usuario
                    print(f"NO SE PUDO AGREGAR AL USUARIO {user_to_add}: {e}")
            else:
                # Si no se encontró el botón después de las 3 iteraciones
                print(
                    f"NO SE PUDO ENCONTRAR EL BOTÓN PARA AGREGAR AL USUARIO {user_to_add}"
                )

        time.sleep(2)

        for a in range(3):
            try:
                cancelar = self.driver.find_element(
                    By.XPATH,
                    f"/html/body/div[{a+2}]/div[2]/div/mat-dialog-container/app-admin-groups/form/mat-dialog-actions/button[1]",
                )

                cancelar.click()
                break
            except (TimeoutException, NoSuchElementException):
                continue
        time.sleep(2)

        # # Varible boton de guardar grupo
        # for a in range(3):
        #     try:
        #         save_group = self.driver.find_element(
        #             By.XPATH,
        #             f"/html/body/div[{a+2}]/div[2]/div/mat-dialog-container/app-admin-groups/form/mat-dialog-actions/button[2]",
        #         )
        #         save_group.click()
        #         break
        #     except (TimeoutException, NoSuchElementException) as e:
        #         print(f"NO SE ENCONTRO EL BOTON 'GUARDAR'. ERROR: {e}")

    # Sección Inicio/Características
    def create_form(
        self,
        form_name,
        group_name,
        campaign_name,
        type_formulario,
        rol_download_list,
        typifi_block_yes_not,
        max_attempts=30,
        delay=1,
    ):
        """Metodo para colocar las caracteristicas del formulario"""

        # Boton ingreso a modulo Fomularios
        button_forms = self.wait.until(
            EC.visibility_of_element_located(
                (
                    By.XPATH,
                    "/html/body/app-root/app-mios/app-side-bar/div/mat-sidenav-container/mat-sidenav/div/app-left-nav/div/div/div/mat-nav-list/div/a[1]",
                )
            )
        )

        button_forms.click()

        # Boton Crear Formulario
        create_button = self.wait.until(
            EC.visibility_of_element_located(
                (
                    By.XPATH,
                    "/html/body/app-root/app-mios/app-side-bar/div/mat-sidenav-container/mat-sidenav-content/div/app-forms-list/div/div/div[2]/div/button",
                )
            )
        )
        create_button.click()

        try:
            # Coloca el nombre del Formulario
            name_form = self.wait.until(
                EC.visibility_of_element_located(
                    (
                        By.XPATH,
                        "/html/body/app-root/app-mios/app-side-bar/div/mat-sidenav-container/mat-sidenav-content/div/app-admin-forms/div/div[1]/div[2]/mat-form-field[1]/div/div[1]/div/input",
                    )
                )
            )
            name_form.send_keys(form_name)
        except (TimeoutException, NoSuchElementException) as e:
            print(f"NO SE PUDO COLOCAR EL NOMBRE DEL FORMULARIO. ERROR :{str(e)}")

        # Elige el tipo del formulario
        type_form = self.wait.until(
            EC.element_to_be_clickable(
                (
                    By.XPATH,
                    "/html/body/app-root/app-mios/app-side-bar/div/mat-sidenav-container/mat-sidenav-content/div/app-admin-forms/div/div[1]/div[2]/mat-form-field[2]/div/div[1]/div/mat-select",
                )
            )
        )
        type_form.click()

        for attempt in range(5):
            try:
                option_type = self.wait.until(
                    EC.element_to_be_clickable(
                        (By.XPATH, f"//mat-option[contains(., '{type_formulario}')]")
                    )
                )
                break
            except (TimeoutException, NoSuchElementException):
                time.sleep(delay)
        if option_type:
            option_type.click()
        else:
            raise ValueError(
                f"NO SE PUDO ENCONTRA EL TIPO DEL FORMULARIO {type_formulario}"
            )

        # Elige que roles descargan
        download_general = self.driver.find_element(
            By.XPATH,
            "/html/body/app-root/app-mios/app-side-bar/div/mat-sidenav-container/mat-sidenav-content/div/app-admin-forms/div/div[1]/div[2]/mat-form-field[3]/div/div[1]/div/mat-select",
        )
        download_general.click()

        # Bucle para Elegir los roles
        for name_rol in rol_download_list:
            for i in range(4):
                try:
                    option = self.wait.until(
                        EC.visibility_of_element_located(
                            (By.XPATH, f"//mat-option[contains(., '{name_rol}')]")
                        )
                    )
                    break
                except (TimeoutException, NoSuchElementException):
                    time.sleep(delay)
            if option:
                option.click()
            else:
                raise ValueError(f"NO SE PUDO ENCONTRAR EL ROL: {name_rol}")

        time.sleep(2)

        download_general.send_keys(Keys.TAB)

        # Elige la campaña a la que pertenece
        campaigns = self.driver.find_element(
            By.XPATH,
            "/html/body/app-root/app-mios/app-side-bar/div/mat-sidenav-container/mat-sidenav-content/div/app-admin-forms/div/div[1]/div[2]/mat-form-field[4]/div/div[1]/div/mat-select",
        )

        campaigns.click()

        # Bucle para buscar la campaña por nombre
        for attempt in range(5):
            try:
                select_campaign = self.wait.until(
                    EC.element_to_be_clickable(
                        (By.XPATH, f"//mat-option[contains(., '{campaign_name}')]")
                    )
                )
                break
            except (TimeoutException, NoSuchElementException):
                time.sleep(delay)
        if select_campaign:
            select_campaign.click()
        else:
            raise ValueError(f"NO SE PUEDE ENCONTAR LA CAMPAÑA: {campaign_name}")

        time.sleep(1)

        # Elige el grupo
        group = self.driver.find_element(
            By.XPATH,
            "/html/body/app-root/app-mios/app-side-bar/div/mat-sidenav-container/mat-sidenav-content/div/app-admin-forms/div/div[1]/div[2]/mat-form-field[5]/div/div[1]",
        )

        group.click()

        # Bucle para buscar el grupo por nombre
        for attempt in range(max_attempts):
            try:
                group_option = self.wait.until(
                    EC.element_to_be_clickable(
                        (By.XPATH, f"//mat-option[contains(., '{group_name}')]")
                    )
                )
                break
            except (TimeoutException, NoSuchElementException):
                time.sleep(delay)
        if group_option:
            group_option.click()
        else:
            raise ValueError(f"NO SE PUDO ENCOTRAR EL GRUPO: {group_name}")

        time.sleep(2)

        # Bucle para seleccionar opciones si y no los dos campos siguientes
        for i, option_yes_not in enumerate(typifi_block_yes_not):

            # Elemento desplegable
            time_of_tipifi = self.driver.find_element(
                By.XPATH,
                f"/html/body/app-root/app-mios/app-side-bar/div/mat-sidenav-container/mat-sidenav-content/div/app-admin-forms/div/div[1]/div[2]/mat-form-field[{i+6}]",
            )

            time_of_tipifi.click()
            time.sleep(1.5)
            if option_yes_not == "si":
                for a in range(3):
                    try:
                        # Opciones si y no
                        option_tipifi1 = self.driver.find_element(
                            By.XPATH,
                            f"/html/body/div[{a+2}]/div[2]/div/div/div/mat-option[1]",
                        )
                        option_tipifi1.click()
                        break
                    except (TimeoutException, NoSuchElementException):
                        continue
            else:
                for a in range(3):
                    try:
                        option_tipifi2 = self.driver.find_element(
                            By.XPATH,
                            f"/html/body/div[{a+2}]/div[2]/div/div/div/mat-option[2]",
                        )
                        option_tipifi2.click()
                        break
                    except (TimeoutException, NoSuchElementException):
                        continue

    # Inicio de la Seccion Accion/Creacion
    def action_create(
        self,
        type_camp,
        name_campo,
        num_colum,
        list_yes_no,
        list_yes_no2,
        name_rol_see_list,
        name_rol_edit_list,
        chacter_min,
        chater_max,
        place_num,
        current_section,
        list_options=None,
        name_section=None,
    ):
        """Metodo Para iniciar la creación del formulario"""

        # Creacion de nueva sección, si el nombre de sección contiene un valor
        if name_section:

            try:
                # Boton para crear la sección
                add_section_button = self.wait.until(
                    EC.element_to_be_clickable(
                        (
                            By.XPATH,
                            "/html/body/app-root/app-mios/app-side-bar/div/mat-sidenav-container/mat-sidenav-content/div/app-admin-forms/div/div[1]/div[3]/div/button",
                        )
                    )
                )
                add_section_button.click()

                # Campo para colocar el nombre de la sección
                section_name_input = self.wait.until(
                    EC.presence_of_element_located(
                        (
                            By.XPATH,
                            f"/html/body/app-root/app-mios/app-side-bar/div/mat-sidenav-container/mat-sidenav-content/div/app-admin-forms/div/div[1]/div[3]/mat-card[{self.current_section}]/div/div[1]/mat-form-field[1]/div/div[1]/div/input",
                        )
                    )
                )
                section_name_input.send_keys(name_section)

                # Campo para elegir el tipo de información
                type_information = self.wait.until(
                    EC.element_to_be_clickable(
                        (
                            By.XPATH,
                            f"/html/body/app-root/app-mios/app-side-bar/div/mat-sidenav-container/mat-sidenav-content/div/app-admin-forms/div/div[1]/div[3]/mat-card[{self.current_section}]/div/div[1]/mat-form-field[2]/div/div[1]/div/mat-select",
                        )
                    )
                )
                type_information.click()

                # Opcion del tipo de información
                option_type = self.wait.until(
                    EC.element_to_be_clickable(
                        (By.XPATH, "//mat-option[contains(., 'Gestión')]")
                    )
                )
                option_type.click()
            except (TimeoutException, NoSuchElementException) as e:
                print(f"NO SE PUDO CREAR LA SECCION. EL ERROR ES :{e}")

        # Tiempo para ubicar el campo
        time.sleep(0.5)

        # Diccionario con los campos y su valor
        campos = {
            "texto": 1,
            "desplegable": 2,
            "multipleseleccion": 3,
            "fecha": 4,
            "agendamiento": 5,
            "numerico": 6,
            "comentario": 7,
            "email": 8,
            "radiobutton": 9,
            "autocomplete": 10,
            "moneda": 11,
            "archivo": 12,
            "tiempo": 13,
        }

        # Almacena el valor por nombre de campo
        num_campo = campos.get(type_camp.lower())

        # Manejo de error en caso de no encotrar el nombre del campo
        if num_campo is None:
            print(f"Tipo de campo no reconocido: {type_camp}")
            print("Tipos de campo válidos:", list(campos.keys()))
            return

        # Base del xpath de los campos
        base_camp_xpath = "/html/body/app-root/app-mios/app-side-bar/div/mat-sidenav-container/mat-sidenav-content/div/app-admin-forms/div/div[2]/mat-tab-group/div/mat-tab-body[1]/div/div/div/div/div"

        # xpath completo del campo
        camp_xpath = f"{base_camp_xpath}[{num_campo}]"

        # Manejo del cambio del xpath para el placeholder cuando se cree una nueva sección
        if self.current_section == 1:
            # Xpath base del placeholder
            base_xpath = "/html/body/app-root/app-mios/app-side-bar/div/mat-sidenav-container/mat-sidenav-content/div/app-admin-forms/div/div[1]/div[3]/mat-card/mat-grid-list/div/mat-grid-tile"
        else:
            # Xpath del placeholder modificado
            base_xpath = f"/html/body/app-root/app-mios/app-side-bar/div/mat-sidenav-container/mat-sidenav-content/div/app-admin-forms/div/div[1]/div[3]/mat-card[{self.current_section}]/mat-grid-list/div/mat-grid-tile"

        # Manejo de cambio de placeholder mientras avanza con los campos
        if self.section_index == 1:
            placeholder_xpath = f"{base_xpath}/figure/div"
        else:
            placeholder_xpath = f"{base_xpath}[{self.section_index}]/figure/div"

        # Elementos campo y placeholder con las modificaciones realizadas
        campo = self.driver.find_element(By.XPATH, camp_xpath)
        placeholder = self.driver.find_element(By.XPATH, placeholder_xpath)

        try:
            # Secuencia de acciones para el movimiento del campo hacia el placeholder
            action = ActionChains(self.driver)
            action.move_to_element(campo)
            action.click_and_hold()
            action.pause(0.5)
            action.move_to_element(placeholder)
            action.move_by_offset(5, 5)
            action.release()
            action.perform()
        except (TimeoutException, NoSuchElementException) as e:
            print(f"NO SE LOGRO COLOCAR EL CAMPO EN SU LUGAR. ERROR: {e}")

        # Tiempo para cargar la parte de variables
        time.sleep(1)

        # Variable del campo nombre del campo
        nombre_campo = self.wait.until(
            EC.visibility_of_element_located(
                (
                    By.XPATH,
                    "/html/body/app-root/app-mios/app-side-bar/div/mat-sidenav-container/mat-sidenav-content/div/app-admin-forms/div/div[2]/mat-tab-group/div/mat-tab-body[2]/div/div/div/mat-form-field[1]/div/div[1]/div/input",
                )
            )
        )
        nombre_campo.send_keys(name_campo)
        time.sleep(1)

        # Variable numero de columnas del campo
        numero_colum = self.wait.until(
            EC.visibility_of_element_located(
                (
                    By.XPATH,
                    "/html/body/app-root/app-mios/app-side-bar/div/mat-sidenav-container/mat-sidenav-content/div/app-admin-forms/div/div[2]/mat-tab-group/div/mat-tab-body[2]/div/div/div/mat-form-field[2]/div/div[1]/div/input",
                )
            )
        )
        numero_colum.clear()
        time.sleep(0.5)
        numero_colum.send_keys(num_colum)

        # Variable para manejar el cambio de xpath de los campos roles que ves e/ editan
        offset = 0

        # Lista de los campos con opciones (Especiales)
        special_types = [
            "desplegable",
            "multipleseleccion",
            "radiobutton",
            "autocomplete",
        ]

        is_special_types = type_camp.lower() in special_types

        # Agrega las opciones si el campo es uno de los especiales lo cual contiene opciones
        if is_special_types and list_options:
            self.add_options(list_options)

        # Bucle para colocar las primeras caracteristicas de los campos
        for i, option_yes_not in enumerate(list_yes_no):
            # Manejo de las caracteristicas para los campos especiales
            xpath_index = i + 2 if is_special_types else i + 1

            if option_yes_not == "si":
                try:
                    option_yes = self.driver.find_element(
                        By.XPATH,
                        f"/html/body/app-root/app-mios/app-side-bar/div/mat-sidenav-container/mat-sidenav-content/div/app-admin-forms/div/div[2]/mat-tab-group/div/mat-tab-body[2]/div/div/div/div[{xpath_index}]/section/mat-radio-group/mat-radio-button[1]",
                    )

                    option_yes.click()
                except (TimeoutException, NoSuchElementException):
                    print(f"NO SE PUDO ENCONTRAR EL ELEMENTO PARA i={i}")
            else:
                try:
                    option_no = self.driver.find_element(
                        By.XPATH,
                        f"/html/body/app-root/app-mios/app-side-bar/div/mat-sidenav-container/mat-sidenav-content/div/app-admin-forms/div/div[2]/mat-tab-group/div/mat-tab-body[2]/div/div/div/div[{xpath_index}]/section/mat-radio-group/mat-radio-button[2]",
                    )
                    option_no.click()
                except (TimeoutException, NoSuchElementException):
                    print(f"NO SE PUDO ENCONTRAR EL ELEMENTO PARA i={i}")
            time.sleep(0.5)
            # print(f"i: {i}")

            # Manejo de cambios en el xpath de rol ven/editan
            if i == 3 and option_yes_not == "si" and not special_types:
                offset = 1
            if i == 2 and option_yes_not == "si" and special_types:
                offset = 1

        #  Variable del desplegable roles que ven
        rol_see = self.driver.find_element(
            By.XPATH,
            f"/html/body/app-root/app-mios/app-side-bar/div/mat-sidenav-container/mat-sidenav-content/div/app-admin-forms/div/div[2]/mat-tab-group/div/mat-tab-body[2]/div/div/div/mat-form-field[{3 + offset}]/div/div[1]/div/mat-select",
        )
        rol_see.click()

        # Bucle para elegir los roles que ven
        for name_rol_see in name_rol_see_list:
            for i in range(1):
                try:
                    rol = self.wait.until(
                        EC.element_to_be_clickable(
                            (By.XPATH, f"//mat-option[contains(., '{name_rol_see}')]")
                        )
                    )
                except (TimeoutException, NoSuchElementException):
                    time.sleep(1)
                if rol:
                    rol.click()
                else:
                    raise ValueError(f"NO SE PUDO ENCONTRAR EL ROL: {name_rol_see}")

        rol_see.send_keys(Keys.TAB)

        # Tiempo necesario para que cargen las opciones del siguiente desplegable
        time.sleep(1.5)

        # Variable de desplegable roles que editan
        rol_edit = self.driver.find_element(
            By.XPATH,
            f"/html/body/app-root/app-mios/app-side-bar/div/mat-sidenav-container/mat-sidenav-content/div/app-admin-forms/div/div[2]/mat-tab-group/div/mat-tab-body[2]/div/div/div/mat-form-field[{4+ offset}]/div/div[1]/div/mat-select",
        )
        rol_edit.click()
        time.sleep(1)

        # Bucle para elegir los roles que editan
        for name_rol_edit in name_rol_edit_list:
            for i in range(1):
                try:
                    rol = self.wait.until(
                        EC.element_to_be_clickable(
                            (By.XPATH, f"//mat-option[contains(., '{name_rol_edit}')]")
                        )
                    )
                except (TimeoutException, NoSuchElementException):
                    time.sleep(1)
            if rol:
                rol.click()
            else:
                raise ValueError(f"NO SE PUDO ENCONTRAR EL ROL: {name_rol_edit}")

        rol_edit.send_keys(Keys.TAB)

        # condicional para la realizacion de los campos maximos/minimos caracteres
        if type_camp not in special_types or ("archivo", "tiempo"):
            try:

                max_chacter = self.driver.find_element(
                    By.XPATH,
                    "/html/body/app-root/app-mios/app-side-bar/div/mat-sidenav-container/mat-sidenav-content/div/app-admin-forms/div/div[2]/mat-tab-group/div/mat-tab-body[2]/div/div/div/div[6]/mat-form-field[2]/div/div[1]/div/input",
                )

                max_chacter.clear()
                max_chacter.send_keys(chater_max)

                min_chacter = self.driver.find_element(
                    By.XPATH,
                    "/html/body/app-root/app-mios/app-side-bar/div/mat-sidenav-container/mat-sidenav-content/div/app-admin-forms/div/div[2]/mat-tab-group/div/mat-tab-body[2]/div/div/div/div[6]/mat-form-field[1]/div/div[1]/div/input",
                )

                min_chacter.send_keys(chacter_min)

            except (TimeoutException, NoSuchElementException) as e:
                print(f"EL ERROR ES: {e}")

        # Manejo de cambios del xpath para las siguientes caracteristicas del campo
        if type_camp in special_types or type_camp in ("archivo", "tiempo"):
            interval = 6
        else:
            interval = 7

        # valor adicionador de los cambios
        additional_offset = 0

        # Condicional para agregar otra opcion que cambia cuando el campos se numerico o moneda
        if type_camp in ("numerico", "moneda"):
            # Posiciones donde se realizan los cambios
            special_positions = [0, 1, 3, 5]
        else:
            # Posiciones donde se realizan los cambios
            special_positions = [0, 1, 5]

        # Bucle para seleccionar las opciones si/no de la segunda lista
        for q, option_y_n2 in enumerate(list_yes_no2):

            current_q = q + additional_offset

            xpath_index = current_q + interval

            if option_y_n2 == "si":
                try:
                    option_yes = self.wait.until(
                        EC.visibility_of_element_located(
                            (
                                By.XPATH,
                                f"/html/body/app-root/app-mios/app-side-bar/div/mat-sidenav-container/mat-sidenav-content/div/app-admin-forms/div/div[2]/mat-tab-group/div/mat-tab-body[2]/div/div/div/div[{xpath_index}]/section/mat-radio-group/mat-radio-button[1]",
                            )
                        )
                    )

                    option_yes.click()
                    option_yes.click()

                    # Manejo de cambio del index si se encuentra en las posiciones especiales
                    if q in special_positions:
                        additional_offset += 1
                except (TimeoutException, NoSuchElementException):
                    print(f"NO SE PUDO ENCONTRAR EL ELEMENTO PARA {current_q}")
            else:
                try:
                    option_no = self.wait.until(
                        EC.visibility_of_element_located(
                            (
                                By.XPATH,
                                f"/html/body/app-root/app-mios/app-side-bar/div/mat-sidenav-container/mat-sidenav-content/div/app-admin-forms/div/div[2]/mat-tab-group/div/mat-tab-body[2]/div/div/div/div[{xpath_index}]/section/mat-radio-group/mat-radio-button[2]",
                            )
                        )
                    )

                    option_no.click()
                    option_no.click()
                except (TimeoutException, NoSuchElementException):
                    print(f"NO SE PUDO ENCONTRAR EL ELEMENTO PARA {current_q}")
            time.sleep(0.5)
            # Impresiones de los index y cambios para indentificarlos
            # print(
            #     f"q original: {q}, q ajustado: {current_q}, índice XPath: {xpath_index}"
            # )
            # print(option_y_n2)
            # print(f"Offset adicional actual: {additional_offset}")

        time.sleep(0.5)

        # Variable del boton guardar campo
        save_camp = self.wait.until(
            EC.visibility_of_element_located(
                (
                    By.XPATH,
                    "/html/body/app-root/app-mios/app-side-bar/div/mat-sidenav-container/mat-sidenav-content/div/app-admin-forms/div/div[2]/mat-tab-group/div/mat-tab-body[2]/div/div/div/button",
                )
            )
        )
        save_camp.click()
        
    # Metodo para la creacion final del formulario
    def finally_save(self,list_camps_filters,list_camps_indentifier):
        """Metodo para la Guardar el formulario creado"""
        
        # Elemento Boton crear formulario
        botton_create_form = self.wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/app-root/app-mios/app-side-bar/div/mat-sidenav-container/mat-sidenav-content/div/app-admin-forms/div/div[1]/div[4]/button[2]")))
        
        botton_create_form.click()
        
        # Elemento desplegable de filtros 
        select_filters = self.wait.until(EC.visibility_of_element_located((By.XPATH, "/html/body/app-root/app-mios/app-side-bar/div/mat-sidenav-container/mat-sidenav-content/div/app-admin-forms/div/div[2]/mat-tab-group/div/mat-tab-body[3]/div/div[1]/mat-form-field/div/div[1]/div/mat-select")))
        
        select_filters.click()
        
        # Bucle para selecionar los campos filtros
        for camps in list_camps_filters:
            for i in range(1):
                try:
                    filters_camp = self.wait.until(EC.element_to_be_clickable((By.XPATH, f"//mat-option[contains(., '{camps}')]"))) 
                except (TimeoutException, NoSuchElementException):
                    time.sleep(1)
            if filters_camp:
                filters_camp.click()
            else:
                raise ValueError(f"NO SE PUDO ENCONTRAR EL CAMPO:{camps}")
        
        # Accion para salir del desplegable  
        select_filters.send_keys(Keys.TAB)
        
        # Elemento boton siguiente
        botton_next = self.wait.until(EC.visibility_of_element_located((By.XPATH, "/html/body/app-root/app-mios/app-side-bar/div/mat-sidenav-container/mat-sidenav-content/div/app-admin-forms/div/div[2]/mat-tab-group/div/mat-tab-body[3]/div/div[3]/div[2]/button")))
        
        # Condicional que verifica que este el boton siguiente
        if botton_next:
            botton_next.click()
        else:
            raise ValueError("NO SE PUDO ENCONTRA EL BOTON: 'SIGUIENTE'")
        
        # Elemento desplegable indetificador unico 
        select_indentifier = self.wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/app-root/app-mios/app-side-bar/div/mat-sidenav-container/mat-sidenav-content/div/app-admin-forms/div/div[2]/mat-tab-group/div/mat-tab-body[4]/div/div[1]/mat-form-field/div/div[1]/div/mat-select")))
        
        select_indentifier.click()
        
        # Bucle para selecionar los campos que sean identifiacador unico
        for camps2 in list_camps_indentifier:
            for i in range(1):
                try:
                    indentifier_camps = self.wait.until(EC.element_to_be_clickable((By.XPATH, f"//mat-option[contains(., '{camps2}')]")))
                except (TimeoutException, NoSuchElementException):
                    time.sleep(1)
            if indentifier_camps:
                
                indentifier_camps.click()
        
        # Accion para salir delm desplegable 
        select_indentifier.send_keys(Keys.TAB)
        
        # Elemento boton Guardar formulario
        save_form = self.wait.until(EC.visibility_of_element_located((By.XPATH, "/html/body/app-root/app-mios/app-side-bar/div/mat-sidenav-container/mat-sidenav-content/div/app-admin-forms/div/div[1]/div[4]/button[2]")))
        
        save_form.click()

    def process_config(self, excel_path):
        """Metodo para almcenar Las primeras variables del Excel"""

        # Variable global que almacena el Excel
        self.excel = load_workbook(excel_path)

        # Variabel con la hoja de Excel config
        excel_config = self.excel["config"]

        # Bucle para sacar las primeras variables de la automatizacion
        for row in excel_config.iter_rows(min_row=2, values_only=True):
            campaing = row[0]
            name_group = row[1]
            descrip_group = row[2]
            sheet_user_name = row[3]
            cell_start = row[4]
            cell_end = row[5]
            form_name = row[6]
            form_type = row[7]
            rol_donwload_list = str(row[8]).split(",") if row[8] else []
            list_y_n = [cell for cell in row[9:11] if cell in ["si", "no"]]
            list_camps_filters = str(row[11]).split(",") if row[11] else []
            indentifier_camps_list = str(row[12]).split(",") if row[12] else []
            # Retornamos las varibles para acceder a ellas
        return (
            campaing,
            name_group,
            descrip_group,
            sheet_user_name,
            cell_start,
            cell_end,
            form_name,
            form_type,
            rol_donwload_list,
            list_y_n,
            list_camps_filters,
            indentifier_camps_list
        )

    # Inicio de la Sección Integracion Excel
    def process_excel(self):
        """Metodo para procesar los datos del excel"""

        # Variable con la hoja de Excel campos
        excel_camps = self.excel["campos"]
        # Variable que contiene la hoja del Excel "opciones"
        excel_options = self.excel["opciones"]

        # Diccionario vacio de opciones
        options_dict = {}
        # Bucle para sacar las opciones del Excel
        for col_index, col in enumerate(
            excel_options.iter_cols(
                min_col=1, max_col=excel_options.max_column, min_row=2, values_only=True
            ),
            start=1,
        ):
            # Bucle que almacena las opciones en listas
            options = [opt for opt in col if opt]
            if options:
                # Agrego de las listas de opciones por index de columna
                options_dict[col_index] = options
        # Contador para los campos especiales que requieren opciones
        special_types_count = 0
        
        # Bucle para tomar los datos las celdas
        for row in excel_camps.iter_rows(min_row=2, values_only=True):
            # Varibles con las caracteristicas de los campos
            type_camp = row[0]
            # Condicional para salir del bucle si ya no hay valores en el excel
            if type_camp is None:
                print("Se han procesado todos los campos del Excel.")
                # Retorna Valor boleano True para Afirmar que ya termino
                return True
            name_campo = row[1]
            num_colum = str(row[2])
            list_yes_no = [cell for cell in row[3:8] if cell in ["si", "no"]]
            list_yes_no2 = [cell for cell in row[8:15] if cell in ["si", "no"]]
            name_rol_see_list = str(row[15]).split(",") if row[15] else []
            name_rol_edit_list = str(row[16]).split(",") if row[16] else []
            chacter_min = str(row[17])
            chater_max = str(row[18])
            new_section = row[19]
            name_section = row[20] if new_section == "si" else None

            special_types = [
                "desplegable",
                "multipleseleccion",
                "radiobutton",
                "autocomplete",
            ]

            # Condicional para aumentar el contador y traer las lista de opciones por numero de contador
            if type_camp.lower() in special_types:
                special_types_count += 1
                list_options = options_dict.get(special_types_count, [])
                if not options:
                    print(f"No se encontro opciones para el campo: {name_campo}")
            else:
                list_options = None

            # Condicional para sumar 1 al section y reiniciar el index si se crea una nueva sección
            if new_section == "si":
                self.current_section += 1
                self.section_index = 1
            else:
                self.section_index += 1

            # Ejecución del método action que toma las variables extraidas del excel
            self.action_create(
                type_camp,
                name_campo,
                num_colum,
                list_yes_no,
                list_yes_no2,
                name_rol_see_list,
                name_rol_edit_list,
                chacter_min,
                chater_max,
                self.section_index,
                self.current_section,
                list_options,
                name_section,
            )


# Metodo main para ejecutar todo el script
def main():
    """Metodo para ejecutar los metodos definidos"""

    try:
        # Instancia de la clase
        Automation = CRM2Automation()

        # Variable que contiene la ruta del archivo Excel
        location_file = input(
            "Por favor, ingrese la ruta completa del archivo Excel: "
        ).strip()

        # Condicional que verfica si el archivo Excel existe
        if not os.path.exists(location_file):
            raise FileNotFoundError(f"El archivo {location_file} no existe.")

        # Trae las varibles del metodo process_config
        (
            campaing,
            name_group,
            descrip_group,
            sheet_user_name,
            cell_start,
            cell_end,
            form_name,
            form_type,
            rol_donwload_list,
            list_y_n,
            list_camps_filters,
            indentifier_camps_list
        ) = Automation.process_config(location_file)

        # Usamos las variables en los demas metodos que lo requieran
        # Ejecución método sección Usuarios por Excel
        users_to_add = Automation.read_user_from_excel(
            location_file, sheet_user_name, cell_start, cell_end
        )

        # Ejecución método sección Inicio/Creación
        Automation.create_group(campaing, name_group, descrip_group, users_to_add)

        # Ejecución método sección Inicio/Características
        Automation.create_form(
            form_name,
            name_group,
            campaing,
            form_type,
            rol_donwload_list,
            list_y_n,
        )

        # Ejecución método sección Logica Integración Excel y Se alamcena el resultado devuelto al terminar
        excel_processed = Automation.process_excel()

        # Condicional para ejecutar el metodo finally_save cuando el metodo process_excel termine
        if excel_processed:
            Automation.finally_save(list_camps_filters,indentifier_camps_list)

    # Manejo de los posibles errores
    except ImportError as e:
        print(f"SE PRODUJO UN ERROR EN: {e}")
    except Exception as e:
        print(f"SE PRODUJO UN ERROR INESPERADO: {e}")


if __name__ == "__main__":
    main()
